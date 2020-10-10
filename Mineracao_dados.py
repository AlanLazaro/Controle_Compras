from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import time

teste = False
cont = 0


wb = load_workbook(filename='planilha.xlsx')
sheetNomes = wb.create_sheet('Nomes')


def abrindoPagina():
    PATH = "chromedriver.exe"
    driver = webdriver.Chrome(PATH)
    return driver

def loginSenha(driver):
    email = driver.find_element_by_id("login")
    senha = driver.find_element_by_id("senha")
    email.send_keys("login")
    senha.send_keys("Senha")
    senha.send_keys(Keys.RETURN)

paginaAberta = abrindoPagina()
paginaAberta.get("-")
loginSenha(paginaAberta)
paginaAberta.get("-ConsultaPedidoCredito.jsp?actualBar=1&pageIndex=0&")

pedidos = []
boxList = paginaAberta.find_element_by_class_name('box_list')

for linhas in boxList.find_elements_by_tag_name('tr'):
    if cont > 0:
        elemento = linhas.find_elements_by_tag_name('td')
        if int(elemento[-3].text) >= 20 and elemento[2].text[3:5] in ('Meses Analisados') and elemento[3].text in ("Liberado para recarga"):
            pedidos.append(elemento[0].text)
    cont += 1

cont = 1
for pedido in pedidos:
    link = "-ConsultaPedidoCreditoItens.jsp?pedido=" + pedido
    paginaAberta.get(link + '&actualBar=1&pageIndex=0&')

    tabela = paginaAberta.find_element_by_class_name('box_list')
    linhas2 = tabela.find_elements_by_tag_name('tr')

    for y in linhas2:
        valoresFinais = y.find_elements_by_tag_name('td')
        try:
            if valoresFinais[4].text == 'Liberado para recarga':
                d = sheetNomes.cell(row=cont, column=1, value=valoresFinais[2].text)
                cont += 1
        except:
            teste = True

wb.save('Planilha Para Salvar Os Nomes')
paginaAberta.close()
